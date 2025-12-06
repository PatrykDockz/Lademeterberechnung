import tkinter as tk
from tkinter import messagebox, ttk, filedialog
import urllib.parse
import urllib.request
import json
import ssl
import openpyxl
import xlrd
from xlutils.copy import copy as xl_copy
import pyperclip
import os

def get_kilometer_von_orten(start, ziel):
    """Berechnet die Entfernung zwischen zwei Orten - vereinfachte Luftlinie"""
    try:
        # SSL-Context für macOS erstellen
        ssl_context = ssl.create_default_context()
        ssl_context.check_hostname = False
        ssl_context.verify_mode = ssl.CERT_NONE
        
        # Geocoding mit Nominatim (OpenStreetMap)
        def get_coords(ort):
            ort_encoded = urllib.parse.quote(ort)
            url = f"https://nominatim.openstreetmap.org/search?q={ort_encoded}&format=json&limit=1"
            req = urllib.request.Request(url, headers={'User-Agent': 'Lademeter-Tool/1.0'})
            with urllib.request.urlopen(req, timeout=30, context=ssl_context) as response:
                data = json.loads(response.read().decode())
                if not data:
                    raise ValueError(f"Ort '{ort}' nicht gefunden")
                return [float(data[0]['lat']), float(data[0]['lon'])]
        
        start_coords = get_coords(start)
        ziel_coords = get_coords(ziel)
        
        # Berechne Luftlinie mit Haversine-Formel (genauer als API bei Timeout-Problemen)
        import math
        lat1, lon1 = math.radians(start_coords[0]), math.radians(start_coords[1])
        lat2, lon2 = math.radians(ziel_coords[0]), math.radians(ziel_coords[1])
        
        dlat = lat2 - lat1
        dlon = lon2 - lon1
        
        a = math.sin(dlat/2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon/2)**2
        c = 2 * math.asin(math.sqrt(a))
        
        # Erdradius in km
        r = 6371
        luftlinie = r * c
        
        # Faktor für Straßenentfernung (ca. 1.3x Luftlinie)
        strassen_km = luftlinie * 1.3
        
        return round(strassen_km, 1)
    except Exception as e:
        raise Exception(f"Fehler bei der Routenberechnung: {str(e)}")

def berechne_lademeter(palettengroesse: str, menge: int, stapelbarkeit: int) -> float:
    try:
        teile = palettengroesse.lower().replace(" ", "").split("x")
        if len(teile) != 3:
            raise ValueError("Ungültiges Format. Bitte verwende das Format 'LxBxH' in cm.")

        laenge_cm = float(teile[0])
        breite_cm = float(teile[1])

        laenge_m = laenge_cm / 100
        breite_m = breite_cm / 100

        lademeter_pro_palette = (breite_m / 2.4) * laenge_m
        gesamt_lademeter = lademeter_pro_palette * menge

        # Stapelbarkeit berücksichtigen
        if stapelbarkeit > 0:
            gesamt_lademeter = gesamt_lademeter / (2 ** stapelbarkeit)
        # stapelbarkeit == 0: keine Änderung

        return round(gesamt_lademeter, 2)

    except Exception as e:
        messagebox.showerror("Fehler", f"Fehler: {e}")
        return 0.0

def berechne_preis(kilometer: float, fahrzeug: str) -> float:
    grundpreis = 65.0
    grund_km = 40
    km_preise = {
        "Sprinter": 0.35,
        "Planensprinter": 0.50,
        "Klein LKW": 0.55,
        "7,5 Tonnen LKW": 0.65,
        "Tautliner": 0.75,
        "Mega": 0.85,
        "Jumbo": 1.00
    }
    km_preis = km_preise.get(fahrzeug, 0.30)
    if kilometer <= grund_km:
        preis = grundpreis
    else:
        preis = grundpreis + (kilometer - grund_km) * km_preis
    return round(preis, 2)

def berechnen():
    palettengroesse = entry_groesse.get()
    try:
        menge = int(entry_menge.get())
        stapelbarkeit = int(entry_stapel.get())
        fahrzeug = combo_fahrzeug.get()
        
        # Prüfe ob Orte oder Kilometer eingegeben wurden
        startort = entry_start.get().strip()
        zielort = entry_ziel.get().strip()
        km_manuell = entry_km.get().strip()
        
        if startort and zielort:
            # Berechne Kilometer automatisch (kostenlos, kein API-Key nötig)
            try:
                label_ergebnis.config(text="🔄 Berechne Route...")
                root.update()
                kilometer = get_kilometer_von_orten(startort, zielort)
                label_ergebnis.config(text=f"📍 Entfernung berechnet: {kilometer} km")
                root.update()
            except Exception as e:
                messagebox.showerror("Fehler", f"Fehler bei der Routenberechnung: {e}\n\nBitte trage die Kilometer manuell ein.")
                return
        elif km_manuell:
            # Verwende manuelle Kilometereingabe
            kilometer = float(km_manuell)
        else:
            messagebox.showerror("Fehler", "Bitte gib entweder Start- und Zielort ODER die Kilometer manuell ein.")
            return
        
        if stapelbarkeit == 0:
            messagebox.showinfo("Hinweis", "Stapelbarkeit 0 bedeutet: Paletten sind nicht stapelbar.")
    except ValueError:
        messagebox.showerror("Fehler", "Bitte gib gültige Zahlen ein.")
        label_ergebnis.config(text="❌ Fehlerhafte Eingabe!")
        return

    lademeter = berechne_lademeter(palettengroesse, menge, stapelbarkeit)
    preis = berechne_preis(kilometer, fahrzeug)
    
    # Speichere Werte global für Tab 2
    global current_lademeter, current_kilometer, current_preis
    current_lademeter = lademeter
    current_kilometer = kilometer
    current_preis = preis
    
    label_ergebnis.config(
        text=f"✅ Die berechneten Lademeter betragen: {lademeter} m\n📏 Entfernung: {kilometer} km\n💶 Ungefährer Preis: {preis} €"
    )

# Excel-Bearbeitungsfunktionen
excel_file_path = None
current_lademeter = None
current_kilometer = None
current_preis = None
is_old_xls = False

def excel_datei_auswaehlen():
    global excel_file_path, is_old_xls
    filepath = filedialog.askopenfilename(
        title="Excel-Datei auswählen",
        filetypes=[("Excel files", "*.xls *.xlsx"), ("All files", "*.*")]
    )
    if filepath:
        excel_file_path = filepath
        is_old_xls = filepath.endswith('.xls')
        excel_label_datei.config(text=f"📄 Datei: {os.path.basename(filepath)}", fg="green")
        messagebox.showinfo("Erfolg", f"Datei geladen:\n{os.path.basename(filepath)}")

def cell_to_index(cell_ref):
    """Konvertiert Excel-Zellreferenz wie 'E14' zu (Zeile, Spalte) für xlrd/xlwt"""
    import re
    match = re.match(r'([A-Z]+)(\d+)', cell_ref)
    if not match:
        return None
    col_str, row_str = match.groups()
    
    # Spalte berechnen (A=0, B=1, etc.)
    col = 0
    for char in col_str:
        col = col * 26 + (ord(char) - ord('A')) + 1
    col -= 1
    
    # Zeile (1-basiert zu 0-basiert)
    row = int(row_str) - 1
    
    return (row, col)

def wolfsburg_einfuegen():
    if not excel_file_path:
        messagebox.showerror("Fehler", "Bitte zuerst eine Excel-Datei auswählen!")
        return
    
    wolfsburg_text = """VW Wolfsburg					
A 39 Ausf. GVZ 					
LKW Steuerstelle Nordstrasse – Anmeldung! 					
KOORDINATES VW 52.436059, 10.750163"""
    
    excel_daten_schreiben('E37', wolfsburg_text)

def excel_daten_schreiben(zelle, wert, max_zeilen=None):
    """
    Schreibt Daten in Excel. Mehrzeilige Texte werden über mehrere Zeilen verteilt.
    
    Args:
        zelle: Startzelle (z.B. 'E14')
        wert: Der zu schreibende Wert (kann mehrzeilig sein)
        max_zeilen: Maximale Anzahl Zeilen (optional)
    """
    if not excel_file_path:
        messagebox.showerror("Fehler", "Bitte zuerst eine Excel-Datei auswählen!")
        return False
    
    try:
        # Prüfe ob mehrzeiliger Text - split nur bei echten Zeilenumbrüchen
        zeilen = []
        if isinstance(wert, str):
            zeilen = wert.split('\n')
            # Entferne leere Zeilen am Ende
            while zeilen and not zeilen[-1].strip():
                zeilen.pop()
        else:
            zeilen = [str(wert)]
        
        if not zeilen:
            zeilen = ['']
        
        if is_old_xls:
            # Alte .xls Datei mit xlrd/xlwt
            rb = xlrd.open_workbook(excel_file_path, formatting_info=True)
            wb = xl_copy(rb)
            
            # Suche nach Sheet "Rechnung"
            sheet_idx = None
            for i in range(rb.nsheets):
                if rb.sheet_by_index(i).name == "Rechnung":
                    sheet_idx = i
                    break
            
            if sheet_idx is None:
                messagebox.showerror("Fehler", "Sheet 'Rechnung' nicht gefunden!")
                return False
            
            ws = wb.get_sheet(sheet_idx)
            start_row, col = cell_to_index(zelle)
            
            # Schreibe jede Zeile in eine separate Zeile
            zeilen_geschrieben = min(len(zeilen), max_zeilen) if max_zeilen else len(zeilen)
            for i in range(zeilen_geschrieben):
                ws.write(start_row + i, col, zeilen[i])
            
            wb.save(excel_file_path)
            messagebox.showinfo("✅ Erfolg", f"Daten wurden ab Zelle {zelle} über {zeilen_geschrieben} Zeile(n) (Sheet: Rechnung) gespeichert!\n\nDatei: {os.path.basename(excel_file_path)}\n\n⚠️ WICHTIG: Bitte schließen Sie die Excel-Datei und öffnen Sie sie neu, um die Änderungen zu sehen!")
            return True
        else:
            # Neue .xlsx Datei mit openpyxl
            wb = openpyxl.load_workbook(excel_file_path)
            
            # Suche nach Sheet "Rechnung"
            if "Rechnung" not in wb.sheetnames:
                messagebox.showerror("Fehler", "Sheet 'Rechnung' nicht gefunden!")
                return False
            
            ws = wb["Rechnung"]
            
            # Hole Start-Zelle
            start_cell = ws[zelle]
            start_row = start_cell.row
            start_col = start_cell.column
            
            # Schreibe jede Zeile in eine separate Zeile
            zeilen_geschrieben = min(len(zeilen), max_zeilen) if max_zeilen else len(zeilen)
            for i in range(zeilen_geschrieben):
                ws.cell(row=start_row + i, column=start_col, value=zeilen[i])
            
            wb.save(excel_file_path)
            messagebox.showinfo("✅ Erfolg", f"Daten wurden ab Zelle {zelle} über {zeilen_geschrieben} Zeile(n) (Sheet: Rechnung) gespeichert!\n\nDatei: {os.path.basename(excel_file_path)}\n\n⚠️ WICHTIG: Bitte schließen Sie die Excel-Datei und öffnen Sie sie neu, um die Änderungen zu sehen!")
            return True
    except Exception as e:
        messagebox.showerror("Fehler", f"Fehler beim Schreiben in {zelle}: {e}")
        return False

def partnerdaten_einfuegen():
    text = excel_entry_e14.get("1.0", "end-1c")
    # E14 bis E19 (Firma, Name, Adresse, PLZ, Ust-IdNr = max 6 Zeilen)
    excel_daten_schreiben('E14', text, max_zeilen=6)

def lade_entlade_einfuegen():
    text = excel_entry_e31_e37.get("1.0", "end-1c")
    zelle = excel_combo_lade_entlade.get()
    if zelle:
        # E31/E37 - Lade-/Entladestelle kann mehrere Zeilen haben
        excel_daten_schreiben(zelle, text)

def fahrzeug_daten_einfuegen():
    text = excel_entry_k42.get("1.0", "end-1c")
    # K42 - Fahrzeug/Paletten/Kilo über mehrere Zeilen
    excel_daten_schreiben('K42', text)

def ids_daten_einfuegen():
    text = excel_entry_k51.get("1.0", "end-1c")
    # K51 - IDs/Dispo über mehrere Zeilen
    excel_daten_schreiben('K51', text)

def kennzeichen_einfuegen():
    text = excel_entry_d22.get().strip()
    excel_daten_schreiben('D22', text)

def fahrername_einfuegen():
    text = excel_entry_j22.get().strip()
    excel_daten_schreiben('J22', text)

def e36_einfuegen():
    text = excel_entry_e36.get().strip()
    excel_daten_schreiben('E36', text)

def e40_einfuegen():
    text = excel_entry_e40.get().strip()
    excel_daten_schreiben('E40', text)

def e35_einfuegen():
    text = excel_entry_e35.get().strip()
    excel_daten_schreiben('E35', text)

def daten_aus_tab1_uebernehmen():
    """Übernimmt Lademeter und Daten aus Tab 1"""
    if current_lademeter and current_kilometer and current_preis:
        text = f"Lademeter: {current_lademeter} m\nKilometer: {current_kilometer} km\nPreis: {current_preis} €"
        excel_entry_k42.delete("1.0", "end")
        excel_entry_k42.insert("1.0", text)
        messagebox.showinfo("Erfolg", "Daten aus Tab 1 übernommen!")
    else:
        messagebox.showwarning("Hinweis", "Bitte zuerst in Tab 1 eine Berechnung durchführen!")

root = tk.Tk()
root.title("Lademeter-Berechnungstool")

# Font-Definitionen
label_font = ("Arial", 12)
entry_font = ("Arial", 12)

# Tab-Control erstellen
tab_control = ttk.Notebook(root)

# Tab 1: Lademeter-Berechnung
tab1 = ttk.Frame(tab_control)
tab_control.add(tab1, text="📦 Lademeter-Berechnung")

# Tab 2: Excel-Bearbeitung
tab2 = ttk.Frame(tab_control)
tab_control.add(tab2, text="📊 Excel-Bearbeitung")

tab_control.pack(expand=1, fill="both")

# ========== TAB 1: Lademeter-Berechnung ==========

tk.Label(tab1, text="Palettengröße (z.B. 120x100x100):", font=label_font).grid(row=0, column=0, sticky="e", padx=10, pady=8)
entry_groesse = tk.Entry(tab1, width=30, font=entry_font)
entry_groesse.grid(row=0, column=1, padx=10, pady=8)

tk.Label(tab1, text="Anzahl der Paletten:", font=label_font).grid(row=1, column=0, sticky="e", padx=10, pady=8)
entry_menge = tk.Entry(tab1, width=30, font=entry_font)
entry_menge.grid(row=1, column=1, padx=10, pady=8)

tk.Label(tab1, text="Stapelbarkeit (z.B. 1, 2, 3):", font=label_font).grid(row=2, column=0, sticky="e", padx=10, pady=8)
entry_stapel = tk.Entry(tab1, width=30, font=entry_font)
entry_stapel.grid(row=2, column=1, padx=10, pady=8)

# Trennlinie für Route/Kilometer Sektion
tk.Label(tab1, text="━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", font=("Arial", 10)).grid(row=3, column=0, columnspan=2, pady=10)

tk.Label(tab1, text="Startort (z.B. Berlin):", font=label_font).grid(row=4, column=0, sticky="e", padx=10, pady=8)
entry_start = tk.Entry(tab1, width=30, font=entry_font)
entry_start.grid(row=4, column=1, padx=10, pady=8)

tk.Label(tab1, text="Zielort (z.B. München):", font=label_font).grid(row=5, column=0, sticky="e", padx=10, pady=8)
entry_ziel = tk.Entry(tab1, width=30, font=entry_font)
entry_ziel.grid(row=5, column=1, padx=10, pady=8)

tk.Label(tab1, text="───── ODER ─────", font=("Arial", 10)).grid(row=6, column=0, columnspan=2, pady=8)

tk.Label(tab1, text="Kilometer (manuell):", font=label_font).grid(row=7, column=0, sticky="e", padx=10, pady=8)
entry_km = tk.Entry(tab1, width=30, font=entry_font)
entry_km.grid(row=7, column=1, padx=10, pady=8)

tk.Label(tab1, text="Fahrzeugtyp:", font=label_font).grid(row=8, column=0, sticky="e", padx=10, pady=8)
combo_fahrzeug = ttk.Combobox(tab1, values=[
    "Sprinter", "Planensprinter", "Klein LKW", "7,5 Tonnen LKW", "Tautliner", "Mega", "Jumbo"
], width=28, font=entry_font, state="readonly")
combo_fahrzeug.current(0)
combo_fahrzeug.grid(row=8, column=1, padx=10, pady=8)

tk.Label(tab1, text="").grid(row=9)  # Leerzeile für Abstand
tk.Label(tab1, text="💡 Tipp: Einfach Start- und Zielort eingeben - Entfernung wird automatisch berechnet!", font=("Arial", 10)).grid(row=10, column=0, columnspan=2, pady=5)
tk.Label(tab1, text="Hinweis: Stapelbarkeit 0 bedeutet, dass Paletten nicht stapelbar sind.", font=("Arial", 10)).grid(row=11, column=0, columnspan=2, pady=3)
tk.Label(tab1, text="Die Berechnung des Preises ist eine Schätzung und kann variieren.", font=("Arial", 10)).grid(row=12, column=0, columnspan=2, pady=3)

btn_berechnen = tk.Button(tab1, text="Berechnen", command=berechnen, bg="#2196F3", fg="white", font=("Arial", 14, "bold"), padx=40, pady=10, cursor="hand2")
btn_berechnen.grid(row=13, column=0, columnspan=2, pady=15)

label_ergebnis = tk.Label(tab1, text="Ergebnis wird hier angezeigt.", font=("Arial", 12), wraplength=500)
label_ergebnis.grid(row=14, column=0, columnspan=2, pady=10)

# ========== TAB 2: Excel-Bearbeitung ==========

# Datei auswählen
tk.Label(tab2, text="Excel-Datei auswählen:", font=("Arial", 14, "bold")).grid(row=0, column=0, columnspan=3, pady=10)
btn_excel_auswaehlen = tk.Button(tab2, text="📁 Datei auswählen", command=excel_datei_auswaehlen, bg="#FF9800", fg="white", font=("Arial", 12, "bold"), padx=20, pady=5)
btn_excel_auswaehlen.grid(row=1, column=0, columnspan=3, pady=5)
excel_label_datei = tk.Label(tab2, text="Keine Datei ausgewählt", font=("Arial", 10), fg="red")
excel_label_datei.grid(row=2, column=0, columnspan=3, pady=5)

# Wolfsburg Button
tk.Label(tab2, text="━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", font=("Arial", 10)).grid(row=3, column=0, columnspan=3, pady=10)
btn_wolfsburg = tk.Button(tab2, text="🚛 Wolfsburg-Daten einfügen (E37)", command=wolfsburg_einfuegen, bg="#4CAF50", fg="white", font=("Arial", 12, "bold"), padx=20, pady=5)
btn_wolfsburg.grid(row=4, column=0, columnspan=3, pady=10)

# Partnerdaten E14
tk.Label(tab2, text="━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", font=("Arial", 10)).grid(row=5, column=0, columnspan=3, pady=10)
tk.Label(tab2, text="Partnerdaten (E14):", font=label_font).grid(row=6, column=0, sticky="w", padx=10, pady=5)
excel_entry_e14 = tk.Text(tab2, width=40, height=3, font=entry_font)
excel_entry_e14.grid(row=7, column=0, columnspan=2, padx=10, pady=5)
btn_e14 = tk.Button(tab2, text="✅ In E14 einfügen", command=partnerdaten_einfuegen, bg="#2196F3", fg="white", font=("Arial", 11, "bold"))
btn_e14.grid(row=7, column=2, padx=5)

# E36, E40, E35
tk.Label(tab2, text="E36:", font=label_font).grid(row=8, column=0, sticky="e", padx=10, pady=5)
excel_entry_e36 = tk.Entry(tab2, width=30, font=entry_font)
excel_entry_e36.grid(row=8, column=1, padx=5)
btn_e36 = tk.Button(tab2, text="✅", command=e36_einfuegen, bg="#2196F3", fg="white", font=("Arial", 11, "bold"))
btn_e36.grid(row=8, column=2, padx=5)

tk.Label(tab2, text="E35:", font=label_font).grid(row=9, column=0, sticky="e", padx=10, pady=5)
excel_entry_e35 = tk.Entry(tab2, width=30, font=entry_font)
excel_entry_e35.grid(row=9, column=1, padx=5)
btn_e35 = tk.Button(tab2, text="✅", command=e35_einfuegen, bg="#2196F3", fg="white", font=("Arial", 11, "bold"))
btn_e35.grid(row=9, column=2, padx=5)

tk.Label(tab2, text="E40:", font=label_font).grid(row=10, column=0, sticky="e", padx=10, pady=5)
excel_entry_e40 = tk.Entry(tab2, width=30, font=entry_font)
excel_entry_e40.grid(row=10, column=1, padx=5)
btn_e40 = tk.Button(tab2, text="✅", command=e40_einfuegen, bg="#2196F3", fg="white", font=("Arial", 11, "bold"))
btn_e40.grid(row=10, column=2, padx=5)

# Lade/Entladestelle E31/E37
tk.Label(tab2, text="━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", font=("Arial", 10)).grid(row=11, column=0, columnspan=3, pady=10)
tk.Label(tab2, text="Lade-/Entladestelle:", font=label_font).grid(row=12, column=0, sticky="w", padx=10, pady=5)
excel_combo_lade_entlade = ttk.Combobox(tab2, values=["E31", "E37"], width=10, font=entry_font, state="readonly")
excel_combo_lade_entlade.current(0)
excel_combo_lade_entlade.grid(row=12, column=1, sticky="w", padx=5)
excel_entry_e31_e37 = tk.Text(tab2, width=40, height=3, font=entry_font)
excel_entry_e31_e37.grid(row=13, column=0, columnspan=2, padx=10, pady=5)
btn_lade = tk.Button(tab2, text="✅ Einfügen", command=lade_entlade_einfuegen, bg="#2196F3", fg="white", font=("Arial", 11, "bold"))
btn_lade.grid(row=13, column=2, padx=5)

# Fahrzeug/Paletten K42
tk.Label(tab2, text="━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", font=("Arial", 10)).grid(row=14, column=0, columnspan=3, pady=10)
tk.Label(tab2, text="Fahrzeug/Paletten/Kilo (K42):", font=label_font).grid(row=15, column=0, sticky="w", padx=10, pady=5)
btn_tab1_uebernehmen = tk.Button(tab2, text="📦 Aus Tab 1 übernehmen", command=daten_aus_tab1_uebernehmen, bg="#9C27B0", fg="white", font=("Arial", 10, "bold"))
btn_tab1_uebernehmen.grid(row=15, column=1, columnspan=2, pady=5)
excel_entry_k42 = tk.Text(tab2, width=40, height=3, font=entry_font)
excel_entry_k42.grid(row=16, column=0, columnspan=2, padx=10, pady=5)
btn_k42 = tk.Button(tab2, text="✅ In K42 einfügen", command=fahrzeug_daten_einfuegen, bg="#2196F3", fg="white", font=("Arial", 11, "bold"))
btn_k42.grid(row=16, column=2, padx=5)

# IDs K51
tk.Label(tab2, text="IDs/Partner Dispo (K51):", font=label_font).grid(row=17, column=0, sticky="w", padx=10, pady=5)
excel_entry_k51 = tk.Text(tab2, width=40, height=3, font=entry_font)
excel_entry_k51.grid(row=18, column=0, columnspan=2, padx=10, pady=5)
btn_k51 = tk.Button(tab2, text="✅ In K51 einfügen", command=ids_daten_einfuegen, bg="#2196F3", fg="white", font=("Arial", 11, "bold"))
btn_k51.grid(row=18, column=2, padx=5)

# Kennzeichen D22
tk.Label(tab2, text="━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", font=("Arial", 10)).grid(row=19, column=0, columnspan=3, pady=10)
tk.Label(tab2, text="Kennzeichen (D22):", font=label_font).grid(row=20, column=0, sticky="e", padx=10, pady=5)
excel_entry_d22 = tk.Entry(tab2, width=30, font=entry_font)
excel_entry_d22.grid(row=20, column=1, padx=5)
btn_d22 = tk.Button(tab2, text="✅", command=kennzeichen_einfuegen, bg="#2196F3", fg="white", font=("Arial", 11, "bold"))
btn_d22.grid(row=20, column=2, padx=5)

# Fahrername J22
tk.Label(tab2, text="Fahrername (J22):", font=label_font).grid(row=21, column=0, sticky="e", padx=10, pady=5)
excel_entry_j22 = tk.Entry(tab2, width=30, font=entry_font)
excel_entry_j22.grid(row=21, column=1, padx=5)
btn_j22 = tk.Button(tab2, text="✅", command=fahrername_einfuegen, bg="#2196F3", fg="white", font=("Arial", 11, "bold"))
btn_j22.grid(row=21, column=2, padx=5)

root.mainloop()
